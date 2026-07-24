#!/usr/bin/env python3
"""Rebuild the MENA dashboard index.html from source files.

Usage:
  python3 rebuild.py --dubai <Dubai.xlsx> --alex <Alexandria.xlsx> --wg <WeGolden.xlsx> \
                     --rev <mena-revenue.csv> --trader <trader-payout.csv> --ip <ip-payout.csv> \
                     [--asof "Jul 31, 2026"] [--out index.html]

CSVs are the UNSUMMARIZED exports of the HubSpot reports:
  "Total MENA Revenue"  -> columns: Create Date, Amount, Deal country, ...
  "MENA Trader Payout"  -> columns: Create date, Approval Amount to Withdraw, Country Ticket, ...
  "MENA IP Payout"      -> same layout as trader payout.
Zip exports are accepted too (the data csv inside is picked automatically).

After rebuilding, verify at least these against the HubSpot summary exports:
total revenue, total payouts, and one spot month. template.html must sit next
to this script. Requires: openpyxl (pip install openpyxl --break-system-packages).
"""
import argparse, csv, datetime, hashlib, io, json, re, sys, warnings, zipfile
from collections import defaultdict
warnings.filterwarnings('ignore')
import openpyxl

MONTHS = {m.lower(): i for i, m in enumerate(
    ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'], 1)}
FULL = {m.lower(): i for i, m in enumerate(
    ['January','February','March','April','May','June','July','August','September','October','November','December'], 1)}
COUNTRY_FIX = {'Palestine, State of': 'Palestine', 'Palestine, State Of': 'Palestine',
               'Syrian Arab Republic': 'Syria', 'syrian arab republic': 'Syria'}

def month_key_from_sheet(name):
    name = name.strip().replace('-', ' ')
    parts = name.split()
    if len(parts) != 2: return None
    mon, yr = parts[0].lower(), parts[1]
    mi = MONTHS.get(mon[:3]) or FULL.get(mon)
    if mi and re.match(r'^\d{4}$', yr):
        return f'{yr}-{mi:02d}'
    return None

def iso_date(v, fallback_mk):
    if isinstance(v, datetime.datetime):
        if v.hour >= 12: v = v + datetime.timedelta(hours=12)
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r'^(\d{1,2})/(\d{1,2})(?:-\d{1,2}/\d{1,2})?/(\d{4})$', s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try: return datetime.date(y, mo, d).strftime('%Y-%m-%d')
            except ValueError: pass
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})(?:T(\d{2}))?', s)
        if m:  # ISO string; UTC evening = next day local (Dubai)
            dt = datetime.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            if m.group(4) and int(m.group(4)) >= 12:
                dt += datetime.timedelta(days=1)
            return dt.strftime('%Y-%m-%d')
    return fallback_mk + '-01' if fallback_mk else None

def load_expenses(dubai, alex, wg):
    txns = []
    # Dubai (AED, signed amounts, header row 2, data from row 3)
    wb = openpyxl.load_workbook(dubai, data_only=True)
    for sn in wb.sheetnames:
        mk = month_key_from_sheet(sn)
        if not mk: continue
        for row in wb[sn].iter_rows(min_row=3, max_col=8):
            no, date, typ, cat, desc, amt, notes, proof = [c.value for c in row]
            if not isinstance(amt, (int, float)): continue
            if not isinstance(no, (int, float)) and not (isinstance(no, str) and no.strip().isdigit()): continue
            txns.append(dict(office='Dubai', mk=mk, date=iso_date(date, mk),
                             type=str(typ or '').strip(), cat=str(cat or '').strip() or 'Uncategorized',
                             desc=str(desc or '').strip(), amt=round(float(amt), 2),
                             cur='AED', notes=str(notes or '').strip()))
    # Alexandria (EGP, expenses positive -> negative; header row 4, data from row 5)
    wb = openpyxl.load_workbook(alex, data_only=True)
    for sn in wb.sheetnames:
        mk = month_key_from_sheet(sn)
        if not mk: continue
        for row in wb[sn].iter_rows(min_row=5, max_col=8):
            no, date, cat, desc, amt, method, notes, receipt = [c.value for c in row]
            first = str(no or '')
            if first.startswith('TOTAL') or first.startswith('Cash Flow') or first.startswith('Carryover'): break
            if not isinstance(amt, (int, float)): continue
            if not isinstance(no, (int, float)) and not (isinstance(no, str) and str(no).strip().isdigit()): continue
            txns.append(dict(office='WMT Alexandria', mk=mk, date=iso_date(date, mk),
                             type='Expense', cat=str(cat or '').strip() or 'Uncategorized',
                             desc=str(desc or '').strip(), amt=-abs(round(float(amt), 2)),
                             cur='EGP', notes=str(notes or '').strip()))
    ws = wb['Transfers']  # funding received, monthly table in cols G/H from row 4
    for row in ws.iter_rows(min_row=4, min_col=7, max_col=8):
        mon, egp = row[0].value, row[1].value
        if not mon or not isinstance(egp, (int, float)) or egp == 0: continue
        mk = month_key_from_sheet(str(mon))
        if not mk: continue
        txns.append(dict(office='WMT Alexandria', mk=mk, date=mk + '-01',
                         type='Transfer In', cat='Funding from HQ',
                         desc='Funds received from Martin (USD->EGP)', amt=round(float(egp), 2),
                         cur='EGP', notes=''))
    # WeGolden (EGP setup block rows 13-27 undated -> Feb 2026; Dubai card AED rows 45-119)
    ws = openpyxl.load_workbook(wg, data_only=True)['WeGolden Egypt eng']
    for row in ws.iter_rows(min_row=13, max_row=27, max_col=6):
        no, typ, cat, desc, amt, notes = [c.value for c in row]
        if not isinstance(amt, (int, float)): continue
        t = str(typ or '').strip()
        sign = 1 if t == 'Income' else -1
        txns.append(dict(office='WeGolden Egypt', mk='2026-02', date='2026-02-01',
                         type='Transfer In' if t == 'Income' else 'Expense',
                         cat=str(cat or '').strip() or 'Uncategorized',
                         desc=str(desc or '').strip(), amt=sign * abs(round(float(amt), 2)),
                         cur='EGP', notes=str(notes or '').strip()))
    last_date = None
    for row in ws.iter_rows(min_row=45, max_row=119, max_col=6):
        no, date, cat, desc, amt, notes = [c.value for c in row]
        if not isinstance(amt, (int, float)): continue
        d = iso_date(date, None) or last_date or '2026-02-01'
        last_date = d
        txns.append(dict(office='WeGolden Egypt', mk=d[:7], date=d,
                         type='Expense', cat=str(cat or '').strip() or 'Uncategorized',
                         desc=(str(desc or '').strip() + ' (Dubai card)'),
                         amt=-round(float(amt), 2), cur='AED', notes=str(notes or '').strip()))
    return txns

def open_csv(path):
    """Accept a .csv or a HubSpot .zip export (picks the non-summary csv)."""
    if str(path).lower().endswith('.zip'):
        z = zipfile.ZipFile(path)
        names = [n for n in z.namelist() if n.endswith('.csv') and 'summary' not in n.lower()]
        if not names: raise SystemExit(f'no data csv inside {path}')
        return io.TextIOWrapper(z.open(names[0]), encoding='utf-8-sig')
    return open(path, encoding='utf-8-sig')

def load_revenue(path):
    rows = list(csv.reader(open_csv(path)))[1:]
    agg = defaultdict(lambda: [0.0, 0])
    for r in rows:
        try: amt = float(r[1])
        except (ValueError, IndexError): continue
        mk, c = r[0][:7], COUNTRY_FIX.get(r[2].strip(), r[2].strip())
        a = agg[(mk, c)]; a[0] += amt; a[1] += 1
    return [{'mk': k[0], 'c': k[1], 'amt': round(v[0], 2), 'n': v[1]} for k, v in sorted(agg.items())]

def load_payouts(trader, ip):
    out = []
    for path, kind in [(trader, 'Trader'), (ip, 'Partner (IP)')]:
        rows = list(csv.reader(open_csv(path)))[1:]
        agg = defaultdict(lambda: [0.0, 0])
        skipped = 0
        for r in rows:
            try: amt = float(r[1])
            except (ValueError, IndexError): skipped += 1; continue
            c = r[2].strip() or 'Unknown'
            if c == '(No value)': c = 'Unknown'
            c = COUNTRY_FIX.get(c, c)
            a = agg[(r[0][:7], c)]; a[0] += amt; a[1] += 1
        if skipped: print(f'  note: {skipped} {kind} row(s) had no amount — skipped', file=sys.stderr)
        out += [{'mk': k[0], 'c': k[1], 'kind': kind, 'amt': round(v[0], 2), 'n': v[1]}
                for k, v in sorted(agg.items())]
    return out

def main():
    ap = argparse.ArgumentParser()
    for a in ['dubai', 'alex', 'wg', 'rev', 'trader', 'ip']: ap.add_argument('--' + a, required=True)
    ap.add_argument('--asof', default=datetime.date.today().strftime('%b %d, %Y'))
    ap.add_argument('--out', default='index.html')
    ap.add_argument('--template', default='template.html')
    args = ap.parse_args()

    txns = load_expenses(args.dubai, args.alex, args.wg)
    rev = load_revenue(args.rev)
    pay = load_payouts(args.trader, args.ip)

    tpl = open(args.template, encoding='utf-8').read()
    out = (tpl.replace('__DATA__', json.dumps(txns, ensure_ascii=False))
              .replace('__REV__', json.dumps(rev))
              .replace('__PAY__', json.dumps(pay))
              .replace('__ASOF__', args.asof))
    open(args.out, 'w', encoding='utf-8').write(out)

    tr = sum(r['amt'] for r in rev); tp = sum(p['amt'] for p in pay)
    print(f'wrote {args.out} ({len(out):,} bytes)')
    print(f'  expenses: {len(txns):,} txns | revenue: ${tr:,.2f} ({sum(r["n"] for r in rev):,} deals) | payouts: ${tp:,.2f}')
    print('verify these totals against the HubSpot summary exports before publishing.')

if __name__ == '__main__':
    main()
